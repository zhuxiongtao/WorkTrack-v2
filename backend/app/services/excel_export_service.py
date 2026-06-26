"""Excel 模块导出服务

按业务模块导出 Excel 表格，支持：
- 主表 + 关联子表分 Sheet 展示
- 外键关联字段展开为可读名称（如 user_id → 申请人姓名）
- 表头加粗 + 冻结首行 + 列宽自适应 + 金额千分位
- 时间范围筛选（created_at）
"""
import io
import logging
from datetime import datetime, date
from typing import Any, Optional
from sqlmodel import Session, select

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.models.user import User
from app.models.department import Department
from app.models.customer import Customer
from app.models.customer_contact import CustomerContact
from app.models.project import Project
from app.models.project_cost import ProjectCost
from app.models.project_follow_up import ProjectFollowUp
from app.models.contract import Contract
from app.models.meeting_note import MeetingNote
from app.models.daily_report import DailyReport
from app.models.weekly_summary import WeeklySummary
from app.models.payment import PaymentRequest
from app.models.seal import SealRequest
from app.models.supplier import Supplier
from app.models.channel import Channel
from app.models.legal_entity import LegalEntity
from app.models.employee_loan import EmployeeLoan
from app.models.expense_request import ExpenseRequest
from app.models.expense_item import ExpenseItem
from app.models.expense_relation import ExpenseRelation
from app.models.business_trip_request import BusinessTripRequest
from app.models.leave_request import LeaveRequest
from app.models.leave_balance import LeaveBalance, LeaveBalanceLog
from app.models.overtime_request import OvertimeRequest
from app.models.purchase_request import PurchaseRequest
from app.models.purchase_supplier import PurchaseSupplier
from app.models.asset import Asset
from app.models.asset_record import AssetRecord
from app.models.rbac import Role, Permission, RolePermission

logger = logging.getLogger("worktrack")


# ──────────────────────────────────────────────────────────────
# 模块字段映射表（核心字段 + 中文列名）
# ──────────────────────────────────────────────────────────────
# 每个模块定义：
#   key: 唯一标识
#   title: 中文模块名（用于 Sheet 名和前端展示）
#   domain: 业务域（core/finance/oa/system）
#   model: SQLModel 类
#   columns: [(字段名, 中文列名), ...]  —— 按导出顺序
#   relations: {字段名: (目标模型, 目标展示字段)}  —— FK 展开
#   sub_sheets: {Sheet名: (子表模型, 外键字段, [(字段名, 中文列名), ...])}
# ──────────────────────────────────────────────────────────────

EXCEL_MODULES: list[dict] = [
    # ═══════════════ 核心业务 ═══════════════
    {
        "key": "customer",
        "title": "客户",
        "domain": "core",
        "model": Customer,
        "columns": [
            ("id", "序号"),
            ("name", "客户名称"),
            ("industry", "行业"),
            ("status", "客户状态"),
            ("contact", "联系方式"),
            ("core_products", "核心产品"),
            ("business_scope", "主营业务"),
            ("scale", "规模"),
            ("website", "官网"),
            ("profile", "公司简介"),
            ("recent_news", "近期动向"),
            ("ai_initiatives", "AI 动向"),
            ("created_at", "创建时间"),
        ],
        "relations": {
            "user_id": (User, "name"),
        },
    },
    {
        "key": "project",
        "title": "项目",
        "domain": "core",
        "model": Project,
        "columns": [
            ("id", "序号"),
            ("customer_name", "客户名称"),
            ("name", "项目名称"),
            ("opportunity_amount", "商机金额"),
            ("opportunity_amount_unit", "商机金额单位"),
            ("deal_amount", "成交金额"),
            ("deal_amount_unit", "成交金额单位"),
            ("currency", "币种"),
            ("start_date", "开始日期"),
            ("termination_date", "终止日期"),
            ("product", "涉及产品"),
            ("project_scenario", "项目场景"),
            ("sales_person", "销售负责人"),
            ("tech_support_person", "技术支持"),
            ("status", "项目状态"),
            ("progress", "进展记录"),
            ("cloud_provider", "客户技术能力"),
            ("cost_amount", "内部成本(元)"),
            ("gross_margin", "毛利率(%)"),
            ("upstream_channels", "上游通道"),
            ("models", "使用模型"),
            ("monthly_call_volume", "月调用量"),
            ("usage_scenario", "使用场景"),
            ("contract_period", "合同周期"),
            ("deadline", "截止日期"),
            ("created_at", "创建时间"),
        ],
        "relations": {
            "user_id": (User, "name"),
            "customer_id": (Customer, "name"),
        },
        "sub_sheets": {
            "项目成本": (ProjectCost, "project_id", [
                ("id", "序号"),
                ("project_id", "项目ID"),
                ("category", "成本类别"),
                ("description", "明细描述"),
                ("amount", "金额"),
                ("cost_month", "发生月份"),
                ("supplier_id", "供应商ID"),
                ("remarks", "备注"),
                ("created_at", "创建时间"),
            ]),
            "项目跟进": (ProjectFollowUp, "project_id", [
                ("id", "序号"),
                ("project_id", "项目ID"),
                ("track", "跟进线"),
                ("content", "跟进内容"),
                ("created_at", "跟进时间"),
            ]),
        },
    },
    {
        "key": "contract",
        "title": "合同",
        "domain": "core",
        "model": Contract,
        "columns": [
            ("id", "序号"),
            ("title", "合同标题"),
            ("contract_no", "合同编号"),
            ("contract_type", "合同类型"),
            ("party_a", "甲方"),
            ("party_b", "乙方"),
            ("contract_amount", "合同金额"),
            ("amount_unit", "金额单位"),
            ("currency", "币种"),
            ("sign_date", "签订日期"),
            ("start_date", "开始日期"),
            ("end_date", "结束日期"),
            ("payment_terms", "付款条件"),
            ("key_clauses", "关键条款"),
            ("summary", "合同摘要"),
            ("status", "合同状态"),
            ("source", "合同来源"),
            ("is_historical", "是否历史归档"),
            ("remarks", "备注"),
            ("created_at", "创建时间"),
        ],
        "relations": {
            "user_id": (User, "name"),
            "customer_id": (Customer, "name"),
            "project_id": (Project, "name"),
        },
    },
    {
        "key": "meeting_note",
        "title": "会议纪要",
        "domain": "core",
        "model": MeetingNote,
        "columns": [
            ("id", "序号"),
            ("title", "会议主题"),
            ("meeting_date", "会议日期"),
            ("attendees", "参会人员"),
            ("content_md", "会议内容"),
            ("ai_summary", "AI摘要"),
            ("audio_url", "录音地址"),
            ("status", "状态"),
            ("created_at", "创建时间"),
        ],
        "relations": {
            "user_id": (User, "name"),
            "customer_id": (Customer, "name"),
            "project_id": (Project, "name"),
        },
    },
    {
        "key": "daily_report",
        "title": "日报",
        "domain": "core",
        "model": DailyReport,
        "columns": [
            ("id", "序号"),
            ("report_date", "日报日期"),
            ("content_md", "工作内容"),
            ("ai_summary", "AI摘要"),
            ("status", "状态"),
            ("created_at", "创建时间"),
        ],
        "relations": {
            "user_id": (User, "name"),
        },
    },
    {
        "key": "weekly_summary",
        "title": "周报",
        "domain": "core",
        "model": WeeklySummary,
        "columns": [
            ("id", "序号"),
            ("week_start", "周开始日期"),
            ("week_end", "周结束日期"),
            ("summary_text", "本周总结"),
            ("status", "状态"),
            ("created_at", "创建时间"),
        ],
        "relations": {
            "user_id": (User, "name"),
        },
    },

    # ═══════════════ 财务 ═══════════════
    {
        "key": "payment_request",
        "title": "付款申请",
        "domain": "finance",
        "model": PaymentRequest,
        "columns": [
            ("id", "序号"),
            ("payment_type", "付款类型"),
            ("title", "摘要"),
            ("amount", "付款金额"),
            ("amount_unit", "金额单位"),
            ("currency", "币种"),
            ("payee", "收款方"),
            ("payee_account", "收款账号"),
            ("reason", "付款事由"),
            ("status", "审批状态"),
            ("created_at", "申请时间"),
        ],
        "relations": {
            "user_id": (User, "name"),
            "contract_id": (Contract, "title"),
        },
    },
    {
        "key": "seal_request",
        "title": "盖章申请",
        "domain": "finance",
        "model": SealRequest,
        "columns": [
            ("id", "序号"),
            ("seal_type", "印章类型"),
            ("title", "用印文件"),
            ("reason", "用印事由"),
            ("copies", "用印份数"),
            ("is_contract_related", "是否涉及合同"),
            ("status", "审批状态"),
            ("created_at", "申请时间"),
        ],
        "relations": {
            "user_id": (User, "name"),
            "contract_id": (Contract, "title"),
        },
    },
    {
        "key": "supplier",
        "title": "供应商",
        "domain": "finance",
        "model": Supplier,
        "columns": [
            ("id", "序号"),
            ("name", "供应商名称"),
            ("code", "简码"),
            ("category", "类型"),
            ("status", "合作状态"),
            ("contact_person", "对接人"),
            ("contact_email", "联系邮箱"),
            ("contact_phone", "联系电话"),
            ("settlement_currency", "结算币种"),
            ("payment_terms", "付款条件"),
            ("contract_start", "合同起始"),
            ("contract_end", "合同终止"),
            ("api_endpoint", "API入口"),
            ("models_provided", "提供模型"),
            ("auth_type", "认证方式"),
            ("total_cost", "累计通道费"),
            ("project_count", "关联项目数"),
            ("remarks", "备注"),
            ("created_at", "创建时间"),
        ],
    },
    {
        "key": "channel",
        "title": "通道",
        "domain": "finance",
        "model": Channel,
        "columns": [
            ("id", "序号"),
            ("name", "通道名称"),
            ("model_type", "模型族"),
            ("kind", "通道类型"),
            ("code", "简码"),
            ("status", "状态"),
            ("cost_price", "成本价"),
            ("price_unit", "计价单位"),
            ("discount_rate", "折扣率"),
            ("suggested_markup", "建议加价率"),
            ("contract_start", "合同起始"),
            ("contract_end", "合同终止"),
            ("inventory_total", "库存总数"),
            ("inventory_available", "在库可用"),
            ("active_projects", "活跃项目数"),
            ("monthly_cost", "当月成本"),
            ("remarks", "备注"),
            ("created_at", "创建时间"),
        ],
        "relations": {
            "supplier_id": (Supplier, "name"),
        },
    },
    {
        "key": "legal_entity",
        "title": "公司主体",
        "domain": "finance",
        "model": LegalEntity,
        "columns": [
            ("id", "序号"),
            ("name", "公司全称"),
            ("short_name", "简称"),
            ("tax_id", "税号"),
            ("balance", "账户余额"),
            ("is_default", "是否默认"),
            ("is_active", "是否启用"),
            ("sort_order", "排序"),
            ("created_at", "创建时间"),
        ],
    },
    {
        "key": "employee_loan",
        "title": "员工借款",
        "domain": "finance",
        "model": EmployeeLoan,
        "columns": [
            ("id", "序号"),
            ("amount", "借款本金"),
            ("used_amount", "已抵消金额"),
            ("remaining", "剩余未还"),
            ("loan_date", "借款日期"),
            ("reason", "借款事由"),
            ("status", "状态"),
            ("created_at", "创建时间"),
        ],
        "relations": {
            "user_id": (User, "name"),
            "entity_id": (LegalEntity, "name"),
        },
    },

    # ═══════════════ OA ═══════════════
    {
        "key": "expense_request",
        "title": "报销申请",
        "domain": "oa",
        "model": ExpenseRequest,
        "columns": [
            ("id", "序号"),
            ("title", "报销摘要"),
            ("expense_type", "费用类型"),
            ("amount", "报销总额"),
            ("amount_unit", "金额单位"),
            ("currency", "币种"),
            ("expense_date", "费用发生时间"),
            ("reason", "报销事由"),
            ("priority_offset_loan", "优先抵消借款"),
            ("offset_loan_amount", "抵扣借款金额"),
            ("company_should_pay", "公司应支付"),
            ("actual_pay_amount", "个人实发"),
            ("company_owes_personal", "公司欠个人"),
            ("status", "审批状态"),
            ("paid_at", "付款时间"),
            ("created_at", "申请时间"),
        ],
        "relations": {
            "user_id": (User, "name"),
            "invoice_entity_id": (LegalEntity, "name"),
            "trip_id": (BusinessTripRequest, "title"),
            "paid_by": (User, "name"),
        },
        "sub_sheets": {
            "报销明细": (ExpenseItem, "expense_id", [
                ("id", "序号"),
                ("expense_id", "报销单ID"),
                ("name", "费用名称"),
                ("expense_type", "费用类型"),
                ("city", "城市"),
                ("expense_date", "费用日期"),
                ("amount", "金额"),
                ("note", "说明"),
                ("remark", "备注"),
                ("sort_order", "排序"),
            ]),
            "关联申请单": (ExpenseRelation, "expense_id", [
                ("id", "序号"),
                ("expense_id", "报销单ID"),
                ("target_type", "关联类型"),
                ("target_id", "关联目标ID"),
                ("relation_note", "关联说明"),
                ("created_at", "创建时间"),
            ]),
        },
    },
    {
        "key": "business_trip",
        "title": "出差申请",
        "domain": "oa",
        "model": BusinessTripRequest,
        "columns": [
            ("id", "序号"),
            ("title", "出差摘要"),
            ("destination", "目的地"),
            ("start_date", "开始日期"),
            ("end_date", "结束日期"),
            ("days", "天数"),
            ("purpose", "出差目的"),
            ("budget", "预算"),
            ("budget_unit", "预算单位"),
            ("currency", "币种"),
            ("transport", "交通方式"),
            ("status", "审批状态"),
            ("completed_at", "完成时间"),
            ("created_at", "申请时间"),
        ],
        "relations": {
            "user_id": (User, "name"),
        },
    },
    {
        "key": "leave_request",
        "title": "请假申请",
        "domain": "oa",
        "model": LeaveRequest,
        "columns": [
            ("id", "序号"),
            ("leave_type", "请假类型"),
            ("title", "请假摘要"),
            ("start_at", "开始时间"),
            ("end_at", "结束时间"),
            ("hours", "请假时长(小时)"),
            ("reason", "请假事由"),
            ("status", "审批状态"),
            ("actual_end_at", "实际销假时间"),
            ("cancelled_at", "销假操作时间"),
            ("created_at", "申请时间"),
        ],
        "relations": {
            "user_id": (User, "name"),
        },
    },
    {
        "key": "overtime_request",
        "title": "加班申请",
        "domain": "oa",
        "model": OvertimeRequest,
        "columns": [
            ("id", "序号"),
            ("title", "加班摘要"),
            ("start_at", "开始时间"),
            ("end_at", "结束时间"),
            ("hours", "加班时长(小时)"),
            ("reason", "加班事由"),
            ("compensate_type", "补偿方式"),
            ("status", "审批状态"),
            ("created_at", "申请时间"),
        ],
        "relations": {
            "user_id": (User, "name"),
        },
    },
    {
        "key": "purchase_request",
        "title": "采购申请",
        "domain": "oa",
        "model": PurchaseRequest,
        "columns": [
            ("id", "序号"),
            ("title", "采购摘要"),
            ("purchase_type", "采购类型"),
            ("total_amount", "总金额"),
            ("amount_unit", "金额单位"),
            ("currency", "币种"),
            ("reason", "采购事由"),
            ("expected_date", "期望到货日期"),
            ("status", "审批状态"),
            ("purchased_at", "采购完成时间"),
            ("stored_at", "入库时间"),
            ("created_at", "申请时间"),
        ],
        "relations": {
            "user_id": (User, "name"),
            "supplier_id": (PurchaseSupplier, "name"),
        },
    },
    {
        "key": "asset",
        "title": "资产",
        "domain": "oa",
        "model": Asset,
        "columns": [
            ("id", "序号"),
            ("name", "资产名称"),
            ("asset_no", "资产编号"),
            ("category", "资产类别"),
            ("spec", "规格型号"),
            ("purchase_date", "购置日期"),
            ("purchase_price", "购置价格"),
            ("amount_unit", "金额单位"),
            ("currency", "币种"),
            ("status", "资产状态"),
            ("location", "存放位置"),
            ("remarks", "备注"),
            ("created_at", "创建时间"),
        ],
        "relations": {
            "user_id": (User, "name"),
            "supplier_id": (PurchaseSupplier, "name"),
        },
        "sub_sheets": {
            "资产流转记录": (AssetRecord, "asset_id", [
                ("id", "序号"),
                ("asset_id", "资产ID"),
                ("action", "操作类型"),
                ("from_status", "变动前状态"),
                ("to_status", "变动后状态"),
                ("from_user_id", "原使用人ID"),
                ("to_user_id", "新使用人ID"),
                ("operator_id", "操作人ID"),
                ("note", "备注"),
                ("created_at", "操作时间"),
            ]),
        },
    },
    {
        "key": "leave_balance",
        "title": "假期余额",
        "domain": "oa",
        "model": LeaveBalance,
        "columns": [
            ("id", "序号"),
            ("leave_type", "假期类型"),
            ("year", "年度"),
            ("total_hours", "总额度(小时)"),
            ("used_hours", "已用额度(小时)"),
            ("created_at", "创建时间"),
        ],
        "relations": {
            "user_id": (User, "name"),
        },
        "sub_sheets": {
            "额度变更日志": (LeaveBalanceLog, "balance_id", [
                ("id", "序号"),
                ("balance_id", "额度账户ID"),
                ("leave_type", "假期类型"),
                ("year", "年度"),
                ("change_type", "变更类型"),
                ("change_hours", "变更时长(小时)"),
                ("reason", "变更原因"),
                ("created_at", "变更时间"),
            ]),
        },
    },
    {
        "key": "purchase_supplier",
        "title": "采购供应商",
        "domain": "oa",
        "model": PurchaseSupplier,
        "columns": [
            ("id", "序号"),
            ("name", "供应商名称"),
            ("short_name", "简称"),
            ("category", "类型"),
            ("status", "合作状态"),
            ("contact_person", "联系人"),
            ("contact_phone", "联系电话"),
            ("contact_email", "联系邮箱"),
            ("address", "地址"),
            ("bank_name", "开户行"),
            ("bank_account", "银行账号"),
            ("tax_no", "纳税人识别号"),
            ("invoice_title", "开票抬头"),
            ("remarks", "备注"),
            ("created_at", "创建时间"),
        ],
    },

    # ═══════════════ 系统 ═══════════════
    {
        "key": "user",
        "title": "用户",
        "domain": "system",
        "model": User,
        "columns": [
            ("id", "序号"),
            ("username", "用户名"),
            ("name", "姓名"),
            ("email", "邮箱"),
            ("is_admin", "是否管理员"),
            ("is_active", "是否启用"),
            ("status", "账号状态"),
            ("job_title", "职位"),
            ("first_work_date", "参加工作日期"),
            ("hire_date", "入职日期"),
            ("last_login_at", "最近登录"),
            ("created_at", "创建时间"),
        ],
        "relations": {
            "department_id": (Department, "name"),
            "leader_id": (User, "name"),
        },
    },
    {
        "key": "department",
        "title": "部门",
        "domain": "system",
        "model": Department,
        "columns": [
            ("id", "序号"),
            ("name", "部门名称"),
            ("description", "描述"),
            ("created_at", "创建时间"),
        ],
    },
    {
        "key": "role",
        "title": "角色权限",
        "domain": "system",
        "model": Role,
        "columns": [
            ("id", "序号"),
            ("name", "角色名称"),
            ("code", "角色编码"),
            ("description", "描述"),
            ("is_system", "是否系统角色"),
            ("created_at", "创建时间"),
        ],
        "sub_sheets": {
            "权限定义": (Permission, None, [
                ("id", "序号"),
                ("code", "权限码"),
                ("name", "权限名称"),
                ("module", "模块"),
                ("action", "操作"),
                ("description", "描述"),
                ("created_at", "创建时间"),
            ]),
            "角色-权限映射": (RolePermission, None, [
                ("id", "序号"),
                ("role_id", "角色ID"),
                ("permission_id", "权限ID"),
            ]),
        },
    },
]

# 按 domain 分组，供前端展示
DOMAIN_LABELS = {
    "core": "核心业务",
    "finance": "财务",
    "oa": "OA 办公",
    "system": "系统管理",
}


def get_modules_summary(db: Session) -> list[dict]:
    """返回可导出模块清单及各模块记录数（供前端展示）"""
    from sqlmodel import func
    result = []
    for mod in EXCEL_MODULES:
        try:
            count = db.exec(select(func.count()).select_from(mod["model"])).one()
        except Exception:
            count = -1
        result.append({
            "key": mod["key"],
            "title": mod["title"],
            "domain": mod["domain"],
            "domain_label": DOMAIN_LABELS.get(mod["domain"], mod["domain"]),
            "count": count,
            "has_sub_sheets": bool(mod.get("sub_sheets")),
        })
    return result


def _format_cell_value(val: Any) -> Any:
    """格式化单元格值"""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, bool):
        return "是" if val else "否"
    return val


def _build_relation_map(db: Session, model, fk_field: str, target_model, target_field: str) -> dict:
    """构建 FK 展开映射：{fk_id: 展示值}"""
    try:
        rows = db.exec(select(target_model)).all()
        return {getattr(r, "id", None): getattr(r, target_field, "") for r in rows}
    except Exception as e:
        logger.warning("构建关联映射 %s.%s → %s.%s 失败: %s", model.__name__, fk_field, target_model.__name__, target_field, e)
        return {}


def _write_sheet(ws, columns: list[tuple], rows: list, relation_maps: dict[str, dict]):
    """向 worksheet 写入数据（表头 + 数据行 + 样式）"""
    # ── 样式定义 ──
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="微软雅黑", size=10)
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # ── 表头 ──
    for col_idx, (_, label) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # ── 数据行 ──
    for row_idx, obj in enumerate(rows, 2):
        for col_idx, (field_name, _) in enumerate(columns, 1):
            # 优先从 relation_map 取展开值
            if field_name in relation_maps:
                raw_val = getattr(obj, field_name, None)
                val = relation_maps[field_name].get(raw_val, raw_val if raw_val else "")
            else:
                val = getattr(obj, field_name, None)
            cell = ws.cell(row=row_idx, column=col_idx, value=_format_cell_value(val))
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border

    # ── 列宽自适应 ──
    for col_idx, (_, label) in enumerate(columns, 1):
        max_len = len(str(label))
        for row_idx in range(2, min(len(rows) + 2, 102)):  # 采样前 100 行
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val:
                max_len = max(max_len, min(len(str(cell_val)), 50))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

    # ── 冻结首行 ──
    ws.freeze_panes = "A2"


def export_excel(
    db: Session,
    module_keys: list[str],
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
) -> io.BytesIO:
    """导出 Excel 到 BytesIO

    Args:
        db: 数据库会话
        module_keys: 要导出的模块 key 列表
        date_from: 起始日期（YYYY-MM-DD），筛选 created_at
        date_to: 截止日期（YYYY-MM-DD），筛选 created_at
    Returns:
        BytesIO 包含 .xlsx 文件
    """
    wb = Workbook()
    # 移除默认 Sheet
    wb.remove(wb.active)

    # 解析日期筛选
    from datetime import datetime as dt
    dt_from = dt.strptime(date_from, "%Y-%m-%d") if date_from else None
    dt_to = dt.strptime(date_to, "%Y-%m-%d").replace(hour=23, minute=59, second=59) if date_to else None

    for mod in EXCEL_MODULES:
        if mod["key"] not in module_keys:
            continue

        model = mod["model"]
        columns = mod["columns"]
        relations = mod.get("relations", {})
        sub_sheets = mod.get("sub_sheets", {})

        # ── 查询主表数据 ──
        stmt = select(model)
        # 时间范围筛选（如果有 created_at 字段）
        if (dt_from or dt_to) and "created_at" in model.model_fields:
            if dt_from:
                stmt = stmt.where(getattr(model, "created_at") >= dt_from)
            if dt_to:
                stmt = stmt.where(getattr(model, "created_at") <= dt_to)
        try:
            rows = db.exec(stmt).all()
        except Exception as e:
            logger.error("查询 %s 失败: %s", mod["title"], e)
            rows = []

        # ── 构建 FK 展开映射 ──
        relation_maps: dict[str, dict] = {}
        for fk_field, (target_model, target_field) in relations.items():
            relation_maps[fk_field] = _build_relation_map(db, model, fk_field, target_model, target_field)

        # ── 写主 Sheet ──
        sheet_name = mod["title"][:31]  # Excel Sheet 名最长 31 字符
        ws = wb.create_sheet(title=sheet_name)
        _write_sheet(ws, columns, rows, relation_maps)

        # ── 写子表 Sheet ──
        for sub_title, (sub_model, sub_fk, sub_columns) in sub_sheets.items():
            try:
                if sub_fk:
                    # 只查询关联到本次主表数据的子记录
                    main_ids = [getattr(r, "id", None) for r in rows if getattr(r, "id", None)]
                    sub_rows = db.exec(select(sub_model).where(getattr(sub_model, sub_fk).in_(main_ids))).all() if main_ids else []
                else:
                    sub_rows = db.exec(select(sub_model)).all()
            except Exception as e:
                logger.error("查询子表 %s 失败: %s", sub_title, e)
                sub_rows = []

            sub_sheet_name = f"{mod['title']}-{sub_title}"[:31]
            ws_sub = wb.create_sheet(title=sub_sheet_name)
            _write_sheet(ws_sub, sub_columns, sub_rows, {})

    # ── 如果没有创建任何 Sheet，加一个空 Sheet ──
    if not wb.sheetnames:
        wb.create_sheet(title="无数据")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
