"""供应商管理 API：MaaS 平台模型供应商 CRUD + 汇总统计 + 关联查询 + 批量导入"""
from collections import defaultdict
from datetime import date
from io import BytesIO
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlmodel import Session, select, func, col
from app.database import get_session
from app.models.supplier import Supplier
from app.models.channel import Channel
from app.models.project_cost import ProjectCost
from app.models.project import Project
from app.schemas.supplier import SupplierCreate, SupplierUpdate, SupplierOut, SupplierSummary
from app.auth import require_permission

router = APIRouter(prefix="/api/v1/suppliers", tags=["供应商管理"])


# ──── 汇总统计（路径固定，必须放在 /{supplier_id} 之前以避免被拦截） ────

@router.get("/summary/all", response_model=list[SupplierSummary])
def get_suppliers_summary(
    db: Session = Depends(get_session),
    current_user=Depends(require_permission("upstream:read")),
):
    """获取所有供应商的汇总统计（成本、项目数、模型列表、本月消费）"""
    current_month = date.today().strftime("%Y-%m")
    suppliers = db.exec(select(Supplier).order_by(col(Supplier.id))).all()
    result = []
    for s in suppliers:
        cost_total = db.exec(
            select(func.coalesce(func.sum(ProjectCost.amount), 0)).where(ProjectCost.supplier_id == s.id)
        ).one()
        project_ids = db.exec(
            select(ProjectCost.project_id).where(ProjectCost.supplier_id == s.id).distinct()
        ).all()
        month_consumed = db.exec(
            select(func.coalesce(func.sum(ProjectCost.amount), 0)).where(
                ProjectCost.supplier_id == s.id,
                ProjectCost.cost_month == current_month,
            )
        ).one()
        result.append(SupplierSummary(
            supplier_id=s.id,
            supplier_name=s.name,
            supplier_code=s.code,
            category=s.category,
            status=s.status,
            settlement_currency=s.settlement_currency,
            total_cost=round(cost_total, 2),
            project_count=len(project_ids),
            models=s.models_provided.split(",") if s.models_provided else [],
            prepaid_balance=s.prepaid_balance,
            current_month_consumed=round(month_consumed, 2),
        ))
    return result


# ──── CRUD ────

@router.get("", response_model=list[SupplierOut])
def list_suppliers(
    status: Optional[str] = None,
    category: Optional[str] = None,
    db: Session = Depends(get_session),
    current_user=Depends(require_permission("upstream:read")),
):
    """获取供应商列表，可按状态/类别筛选"""
    query = select(Supplier).order_by(col(Supplier.id))
    if status:
        query = query.where(Supplier.status == status)
    if category:
        query = query.where(Supplier.category == category)
    return db.exec(query).all()


@router.post("", response_model=SupplierOut)
def create_supplier(
    body: SupplierCreate,
    db: Session = Depends(get_session),
    current_user=Depends(require_permission("upstream:edit")),
):
    """新增供应商，直接生效"""
    existing = db.exec(select(Supplier).where(Supplier.name == body.name)).first()
    if existing:
        raise HTTPException(400, f"供应商 '{body.name}' 已存在")
    obj = Supplier(**body.model_dump())
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj


@router.get("/{supplier_id}", response_model=SupplierOut)
def get_supplier(
    supplier_id: int,
    db: Session = Depends(get_session),
    current_user=Depends(require_permission("upstream:read")),
):
    supplier = db.get(Supplier, supplier_id)
    if not supplier:
        raise HTTPException(404, "供应商不存在")
    return supplier


@router.put("/{supplier_id}", response_model=SupplierOut)
def update_supplier(
    supplier_id: int,
    body: SupplierUpdate,
    db: Session = Depends(get_session),
    current_user=Depends(require_permission("upstream:edit")),
):
    """更新供应商信息"""
    supplier = db.get(Supplier, supplier_id)
    if not supplier:
        raise HTTPException(404, "供应商不存在")
    data = body.model_dump(exclude_unset=True)
    # 如果改了 name，检查唯一性
    if "name" in data and data["name"] != supplier.name:
        dup = db.exec(select(Supplier).where(Supplier.name == data["name"])).first()
        if dup:
            raise HTTPException(400, f"供应商 '{data['name']}' 已存在")
    for k, v in data.items():
        setattr(supplier, k, v)
    db.add(supplier)
    db.commit()
    db.refresh(supplier)
    return supplier


@router.delete("/{supplier_id}")
def delete_supplier(
    supplier_id: int,
    db: Session = Depends(get_session),
    current_user=Depends(require_permission("upstream:edit")),
):
    """删除供应商（需先解除关联的成本条目与通道）"""
    supplier = db.get(Supplier, supplier_id)
    if not supplier:
        raise HTTPException(404, "供应商不存在")
    # 检查是否有关联成本条目
    linked_cost = db.exec(
        select(func.count()).where(ProjectCost.supplier_id == supplier_id)
    ).one()
    if linked_cost > 0:
        raise HTTPException(400, f"该供应商下有 {linked_cost} 条成本记录，请先解除关联后再删除")
    # 检查是否有关联通道
    linked_channel = db.exec(
        select(func.count()).where(Channel.supplier_id == supplier_id)
    ).one()
    if linked_channel > 0:
        raise HTTPException(400, f"该供应商下有 {linked_channel} 个通道，请先删除或调整后再删除供应商")
    db.delete(supplier)
    db.commit()
    return {"ok": True}


class BatchDeleteRequest(BaseModel):
    ids: list[int]


@router.post("/batch-delete")
def batch_delete_suppliers(
    body: BatchDeleteRequest,
    db: Session = Depends(get_session),
    current_user=Depends(require_permission("upstream:edit")),
):
    """批量删除供应商，逐个校验关联成本/通道，部分失败不影响其余项"""
    if not body.ids:
        raise HTTPException(400, "未选择任何供应商")
    if len(body.ids) > 200:
        raise HTTPException(400, "单次最多删除 200 个供应商")

    smap = {s.id: s for s in db.exec(select(Supplier).where(col(Supplier.id).in_(body.ids))).all()}
    deleted: list[int] = []
    failed: list[dict] = []
    for sid in body.ids:
        s = smap.get(sid)
        if not s:
            failed.append({"id": sid, "name": None, "reason": "供应商不存在"})
            continue
        linked_cost = db.exec(select(func.count()).where(ProjectCost.supplier_id == sid)).one()
        if linked_cost > 0:
            failed.append({"id": sid, "name": s.name, "reason": f"有 {linked_cost} 条成本记录关联"})
            continue
        linked_channel = db.exec(select(func.count()).where(Channel.supplier_id == sid)).one()
        if linked_channel > 0:
            failed.append({"id": sid, "name": s.name, "reason": f"有 {linked_channel} 个通道关联"})
            continue
        db.delete(s)
        deleted.append(sid)
    db.commit()
    return {"deleted": deleted, "failed": failed}


# ──── 关联查询 ────

@router.get("/{supplier_id}/projects")
def get_supplier_projects(
    supplier_id: int,
    db: Session = Depends(get_session),
    current_user=Depends(require_permission("upstream:read")),
):
    """获取供应商关联的项目列表及成本明细"""
    supplier = db.get(Supplier, supplier_id)
    if not supplier:
        raise HTTPException(404, "供应商不存在")

    # 查找该供应商关联的成本条目
    cost_items = db.exec(
        select(ProjectCost).where(ProjectCost.supplier_id == supplier_id).order_by(ProjectCost.cost_month)
    ).all()

    # 按项目分组
    project_cost_map: dict[int, list] = defaultdict(list)
    for item in cost_items:
        project_cost_map[item.project_id].append(item)

    # 构建项目列表
    projects_data = []
    for pid, items in project_cost_map.items():
        project = db.get(Project, pid)
        if not project:
            continue
        total_cost = sum(i.amount for i in items)
        # 计算该项目毛利率
        deal = project.deal_amount or 0
        gross_profit = (deal - total_cost) if deal else None
        gross_margin = round((1 - total_cost / deal) * 100, 2) if deal > 0 else None
        projects_data.append({
            "project_id": pid,
            "project_name": project.name,
            "customer_name": project.customer_name,
            "currency": project.currency,
            "deal_amount": project.deal_amount,
            "status": project.status,
            "sales_person": project.sales_person,
            "total_cost": round(total_cost, 2),
            "gross_profit": round(gross_profit, 2) if gross_profit is not None else None,
            "gross_margin": gross_margin,
            "cost_count": len(items),
            "cost_items": [
                {
                    "id": i.id,
                    "category": i.category,
                    "description": i.description,
                    "amount": i.amount,
                    "cost_month": i.cost_month,
                    "remarks": i.remarks,
                }
                for i in items
            ],
        })

    # 按毛利率升序（低毛利排前，提醒关注）
    projects_data.sort(key=lambda x: (x["gross_margin"] if x["gross_margin"] is not None else 999, -x["total_cost"]))

    return {
        "supplier": {
            "id": supplier.id,
            "name": supplier.name,
            "code": supplier.code,
            "category": supplier.category,
            "status": supplier.status,
            "settlement_currency": supplier.settlement_currency,
        },
        "projects": projects_data,
        "total_cost": round(sum(i.amount for i in cost_items), 2),
        "total_projects": len(projects_data),
    }


@router.post("/{supplier_id}/sync-stats")
def sync_supplier_stats(
    supplier_id: int,
    db: Session = Depends(get_session),
    current_user=Depends(require_permission("upstream:edit")),
):
    """手动同步供应商的业务统计数据（累计成本、项目数、本月已消费）"""
    from datetime import date
    supplier = db.get(Supplier, supplier_id)
    if not supplier:
        raise HTTPException(404, "供应商不存在")

    cost_total = db.exec(
        select(func.coalesce(func.sum(ProjectCost.amount), 0)).where(ProjectCost.supplier_id == supplier_id)
    ).one()
    project_ids = db.exec(
        select(ProjectCost.project_id).where(ProjectCost.supplier_id == supplier_id).distinct()
    ).all()

    current_month = date.today().strftime("%Y-%m")
    month_consumed = db.exec(
        select(func.coalesce(func.sum(ProjectCost.amount), 0)).where(
            ProjectCost.supplier_id == supplier_id,
            ProjectCost.cost_month == current_month,
        )
    ).one()

    supplier.total_cost = round(cost_total, 2)
    supplier.project_count = len(project_ids)
    supplier.current_month_consumed = round(month_consumed, 2)
    db.add(supplier)
    db.commit()
    db.refresh(supplier)
    return {
        "ok": True,
        "total_cost": supplier.total_cost,
        "project_count": supplier.project_count,
        "current_month_consumed": supplier.current_month_consumed,
    }




# ──── 批量导入 ────

# 列头映射：Excel 列名 → 内部字段（供应商管理 sheet）
_SUPPLIER_COL_MAP = {
    "供应商名称": "name",
    "简称/编码": "code",
    "供应商类型": "category",
    "合作状态": "status",
    "结算方式": "settlement_method",
    "结算周期(天)": "settlement_cycle_days",
    "预付余额(元)": "prepaid_balance",
    "信用额度(元)": "credit_limit",
    "联系人": "contact_person",
    "联系电话": "contact_phone",
    "联系邮箱": "contact_email",
    "微信/飞书群": "im_group",
    "合同起始日": "contract_start",
    "合同到期日": "contract_end",
    "API文档地址": "api_doc_url",
    "备注": "remarks",
}

_CHANNEL_COL_MAP = {
    "供应模型": "model_family",
    "模型名称": "name",
    "实际结算折扣": "cost_discount",
    "相关更新": "remarks",
}


def _parse_supplier_sheet(df) -> tuple[list[dict], list[dict]]:
    """解析供应商 sheet，返回 (supplier_rows, channel_rows)。每行 channel 附带 _supplier_name。"""
    import pandas as pd
    supplier_rows: list[dict] = []
    channel_rows: list[dict] = []
    seen_suppliers: set[str] = set()

    # 线下表格常将「供应商名称」列合并单元格以表示同一供应商多个通道；
    # pandas 读取合并区域时，除首行外其余行该列为空，需向下填充，
    # 否则这些行会因识别不到供应商名称被整行跳过，导致通道被静默丢失。
    if "供应商名称" in df.columns:
        df = df.copy()
        df["供应商名称"] = df["供应商名称"].ffill()

    for _, row in df.iterrows():
        raw = {k: (None if (isinstance(v, float) and __import__('math').isnan(v)) else v)
               for k, v in row.items()}

        supplier_name = str(raw.get("供应商名称") or "").strip()
        if not supplier_name:
            continue

        # 供应商行（每个名称只取第一次出现的账户字段）
        if supplier_name not in seen_suppliers:
            seen_suppliers.add(supplier_name)
            s_data: dict = {"name": supplier_name}
            for col_name, field in _SUPPLIER_COL_MAP.items():
                if field == "name":
                    continue
                val = raw.get(col_name)
                if val is not None and str(val).strip():
                    if field in ("settlement_cycle_days",):
                        try:
                            val = int(val)
                        except (ValueError, TypeError):
                            val = None
                    elif field in ("prepaid_balance", "credit_limit"):
                        try:
                            val = float(val)
                        except (ValueError, TypeError):
                            val = None
                    else:
                        val = str(val).strip()
                    if val is not None:
                        s_data[field] = val
            supplier_rows.append(s_data)

        # 通道行
        c_data: dict = {"_supplier_name": supplier_name}
        for col_name, field in _CHANNEL_COL_MAP.items():
            val = raw.get(col_name)
            if val is not None and str(val).strip():
                if field == "cost_discount":
                    try:
                        val = float(str(val).replace("折", "").strip())
                        if val > 1:
                            val = val / 100  # "75" → 0.75
                    except (ValueError, TypeError):
                        val = None
                else:
                    val = str(val).strip()
                if val is not None:
                    c_data[field] = val
        if c_data.get("model_family") or c_data.get("name"):
            channel_rows.append(c_data)

    return supplier_rows, channel_rows


@router.get("/import/template")
def download_import_template(
    current_user=Depends(require_permission("upstream:read")),
):
    """下载供应商导入模板（与线下 Excel 列头一致）"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "供应商管理"

    headers = [
        "供应商名称", "供应模型", "模型名称", "简称/编码", "供应商类型", "合作状态",
        "结算方式", "结算周期(天)", "预付余额(元)", "信用额度(元)",
        "实际结算折扣",
        "联系人", "联系电话", "联系邮箱", "微信/飞书群",
        "合同起始日", "合同到期日", "API文档地址", "备注",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", start_color="D9E1F2")
    header_font = Font(bold=True, name="Arial")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 18

    # 示例行
    ws.append(["QY", "Gemini系列", "gemini-2.5-pro", "QY", "渠道代理", "合作中",
               "月结", 30, "", "", "0.75",
               "张三", "138xxxxxxxx", "contact@example.com", "xx群",
               "2026-01", "2027-01", "https://docs.example.com", ""])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=supplier_import_template.xlsx"},
    )


@router.post("/import")
async def import_suppliers(
    file: UploadFile = File(...),
    dry_run: bool = Query(False),
    db: Session = Depends(get_session),
    current_user=Depends(require_permission("upstream:edit")),
):
    """批量导入供应商 + 通道（供应商管理 sheet）。dry_run=true 仅预检不写库。

    upsert 键：供应商按 name，通道按 (supplier_id, name)。
    """
    import pandas as pd

    content = await file.read()
    try:
        df = pd.read_excel(BytesIO(content), sheet_name="供应商管理")
    except Exception as e:
        raise HTTPException(400, f"读取 Excel 失败：{e}")

    supplier_rows, channel_rows = _parse_supplier_sheet(df)

    stats = {"suppliers": {"create": 0, "update": 0}, "channels": {"create": 0, "update": 0}}
    errors: list[dict] = []

    # ── 处理供应商 ──
    supplier_name_to_id: dict[str, int] = {}
    for idx, s_data in enumerate(supplier_rows):
        name = s_data["name"]
        try:
            existing = db.exec(select(Supplier).where(Supplier.name == name)).first()
            if existing:
                for k, v in s_data.items():
                    if k != "name" and v is not None:
                        setattr(existing, k, v)
                if not dry_run:
                    db.add(existing)
                supplier_name_to_id[name] = existing.id if existing.id else -1
                stats["suppliers"]["update"] += 1
            else:
                new_s = Supplier(**{k: v for k, v in s_data.items() if v is not None})
                if not dry_run:
                    db.add(new_s)
                    db.flush()
                    supplier_name_to_id[name] = new_s.id
                else:
                    supplier_name_to_id[name] = -1
                stats["suppliers"]["create"] += 1
        except Exception as e:
            errors.append({"sheet": "供应商管理", "row": idx + 2, "supplier": name, "reason": str(e)})

    if not dry_run:
        db.commit()

    # ── 处理通道 ──
    for idx, c_data in enumerate(channel_rows):
        supplier_name = c_data.pop("_supplier_name")
        supplier_id = supplier_name_to_id.get(supplier_name)
        if supplier_id is None:
            errors.append({"sheet": "供应商管理", "row": idx + 2, "reason": f"找不到供应商「{supplier_name}」"})
            continue
        if dry_run and supplier_id == -1:
            stats["channels"]["create"] += 1
            continue

        model_family = c_data.get("model_family", "")
        channel_name = c_data.get("name") or model_family or "未命名"
        try:
            existing_ch = db.exec(
                select(Channel).where(
                    Channel.supplier_id == supplier_id,
                    Channel.name == channel_name,
                )
            ).first()
            if existing_ch:
                for k, v in c_data.items():
                    if v is not None:
                        setattr(existing_ch, k, v)
                if not dry_run:
                    db.add(existing_ch)
                stats["channels"]["update"] += 1
            else:
                new_ch = Channel(
                    supplier_id=supplier_id,
                    name=channel_name,
                    **{k: v for k, v in c_data.items() if k not in ("name",) and v is not None},
                )
                if not dry_run:
                    db.add(new_ch)
                stats["channels"]["create"] += 1
        except Exception as e:
            errors.append({"sheet": "供应商管理", "row": idx + 2, "supplier": supplier_name, "reason": str(e)})

    if not dry_run:
        db.commit()

    return {
        "dry_run": dry_run,
        "stats": stats,
        "errors": errors,
        "message": "预检完成，未写入数据库" if dry_run else "导入完成",
    }
