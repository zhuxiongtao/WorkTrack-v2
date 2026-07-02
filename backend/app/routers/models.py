"""模型目录相关 API
- GET    /api/v1/models                 业务侧消费：只返回 is_active=true（按 region/provider 过滤）
- GET    /api/v1/models/all             管理后台：返回所有（包含待审校）
- PATCH  /api/v1/models/{id}           审校（编辑/启用/停用）
- DELETE /api/v1/models/{id}           删除
- POST   /api/v1/models/refresh        手动触发刷新
- GET    /api/v1/models/refresh/status 查看上次刷新状态 + 下次定时
"""
import json
import logging
import time
from datetime import datetime, timezone
from app.utils.time import BEIJING_TZ, now
from typing import Optional

from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select, or_, and_

from app.database import engine, get_session
from app.models.model_catalog import ModelCatalog
from app.models.user import User
from app.schemas.model_catalog import (
    ModelCatalogOut,
    ModelCatalogUpdate,
    ModelCatalogListItem,
    ModelRefreshStatus,
    ModelRefreshTriggerResult,
)
from app.auth import get_current_user
from app.services.model_catalog_fetcher import (
    refresh_and_record,
    get_last_status,
    infer_region_from_text,
    _normalize_region,
)
from app.services.scheduler import scheduler
from app.config import settings
import os

router = APIRouter(prefix="/api/v1/models", tags=["models"])
logger = logging.getLogger("worktrack.models_api")

# 手动刷新节流：同 IP/用户 1 小时内最多 1 次
_last_manual_trigger: dict = {}
MANUAL_TRIGGER_COOLDOWN = 3600  # 秒


def _is_admin(user: User) -> bool:
    return bool(getattr(user, "is_admin", False) or getattr(user, "is_superuser", False))


@router.get("", response_model=list[ModelCatalogListItem])
def list_active_models(
    region: Optional[str] = Query(None, description="domestic / international"),
    provider: Optional[str] = Query(None),
    _user: User = Depends(get_current_user),
):
    """业务侧消费：只返回已审校（is_active=true）的模型"""
    with Session(engine) as db:
        stmt = select(ModelCatalog).where(ModelCatalog.is_active == True)
        if region:
            stmt = stmt.where(ModelCatalog.region == region)
        if provider:
            stmt = stmt.where(ModelCatalog.provider == provider)
        stmt = stmt.order_by(ModelCatalog.region.asc(), ModelCatalog.provider.asc(), ModelCatalog.name.asc())
        rows = db.exec(stmt).all()
    return [ModelCatalogListItem.model_validate(r) for r in rows]


@router.get("/all", response_model=list[ModelCatalogOut])
def list_all_models(
    include_inactive: bool = True,
    _user: User = Depends(get_current_user),
):
    """管理后台：返回所有（含待审校 is_active=false）"""
    if not _is_admin(_user):
        raise HTTPException(status_code=403, detail="仅管理员可访问")
    with Session(engine) as db:
        stmt = select(ModelCatalog)
        if not include_inactive:
            stmt = stmt.where(ModelCatalog.is_active == True)
        stmt = stmt.order_by(
            ModelCatalog.is_active.desc(),
            ModelCatalog.last_seen_at.desc().nullslast(),
        )
        rows = db.exec(stmt).all()
    return [ModelCatalogOut.model_validate(r) for r in rows]


@router.patch("/{model_id}", response_model=ModelCatalogOut)
def update_model(
    model_id: int,
    payload: ModelCatalogUpdate,
    user: User = Depends(get_current_user),
):
    """管理员审校：编辑字段 / 启用 / 停用"""
    if not _is_admin(user):
        raise HTTPException(status_code=403, detail="仅管理员可操作")
    with Session(engine) as db:
        row = db.get(ModelCatalog, model_id)
        if not row:
            raise HTTPException(status_code=404, detail="模型不存在")
        data = payload.model_dump(exclude_unset=True)
        for k, v in data.items():
            setattr(row, k, v)
        if "is_active" in data and data["is_active"]:
            row.reviewed_at = now()
            row.reviewed_by = user.id
        row.updated_at = now()
        db.add(row)
        db.commit()
        db.refresh(row)
    return ModelCatalogOut.model_validate(row)


@router.delete("/{model_id}")
def delete_model(
    model_id: int,
    user: User = Depends(get_current_user),
):
    if not _is_admin(user):
        raise HTTPException(status_code=403, detail="仅管理员可操作")
    with Session(engine) as db:
        row = db.get(ModelCatalog, model_id)
        if not row:
            raise HTTPException(status_code=404, detail="模型不存在")
        db.delete(row)
        db.commit()
    return {"ok": True}


@router.post("/refresh", response_model=ModelRefreshTriggerResult)
def manual_refresh(
    user: User = Depends(get_current_user),
):
    """管理员手动触发刷新（带 1 小时节流）"""
    if not _is_admin(user):
        raise HTTPException(status_code=403, detail="仅管理员可操作")
    # 节流
    now = time.time()
    last = _last_manual_trigger.get(user.id, 0)
    if now - last < MANUAL_TRIGGER_COOLDOWN:
        remain = int(MANUAL_TRIGGER_COOLDOWN - (now - last))
        raise HTTPException(status_code=429, detail=f"刷新冷却中，请 {remain // 60} 分钟后再试")
    _last_manual_trigger[user.id] = now
    # 调用
    res = refresh_and_record()
    if not res.get("success"):
        raise HTTPException(status_code=500, detail=res.get("error", "刷新失败"))
    return ModelRefreshTriggerResult(
        success=True,
        inserted=res.get("inserted", 0),
        updated=res.get("updated", 0),
        skipped=0,
        translated=res.get("translated", 0),
        duration_ms=res.get("duration_ms", 0),
    )


@router.get("/refresh/status", response_model=ModelRefreshStatus)
def refresh_status(
    _user: User = Depends(get_current_user),
):
    """查看上次刷新状态 + 下次定时"""
    last = get_last_status()
    enabled_env = os.getenv("MODEL_REFRESH_ENABLED", "true").lower() != "false"
    cron = os.getenv("MODEL_REFRESH_CRON", "0 3 * * 1")
    next_run = None
    try:
        job = scheduler.get_job("model_catalog_refresh")
        if job:
            next_run = job.next_run_time
    except Exception:
        pass
    finished_at = last.get("finished_at")
    if finished_at:
        try:
            finished_at_dt = datetime.fromisoformat(finished_at.replace("Z", "+00:00"))
        except Exception:
            finished_at_dt = None
    else:
        finished_at_dt = None
    return ModelRefreshStatus(
        last_refresh_at=finished_at_dt,
        last_refresh_status="success" if last.get("success") else ("failed" if last.get("error") else None),
        last_refresh_count=(last.get("inserted", 0) + last.get("updated", 0)),
        last_error=last.get("error"),
        next_run_at=next_run,
        enabled=enabled_env,
        cron=cron,
    )


# ──── 模型基础列表导入 ────

# 列头映射：Excel 列名 → 内部字段
_MODEL_COL_MAP = {
    "模型厂商": "provider",
    "模型名称": "name",
    "阶梯计价": "_tier_label",
    "输入价格": "input_price",
    "输出价格": "output_price",
    "价格币种": "price_currency",
    "价格单位": "price_unit",
    "模型状态": "_status",
    "供应商": "suppliers_list",
    "备注": "description",
}

_STATUS_MAP = {"可使用": True, "不可用": False, "已下线": False, "未激活": False, "停用": False}


@router.get("/import/template")
def download_model_import_template(
    _user: User = Depends(get_current_user),
):
    """下载模型基础列表导入模板（与线下 Excel 列头一致）"""
    if not _is_admin(_user):
        raise HTTPException(status_code=403, detail="仅管理员可操作")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "模型基础列表"
    headers = ["模型厂商", "模型名称", "地域", "阶梯计价", "输入价格", "输出价格", "价格币种", "价格单位", "模型状态", "供应商", "备注"]
    ws.append(headers)
    header_fill = PatternFill("solid", start_color="D9E1F2")
    header_font = Font(bold=True, name="Arial")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 20
    ws.append(["Anthropic", "claude-sonnet-4-6", "", "-", 3, 15, "USD", "美元/百万tokens", "可使用", "代理商1、代理商7", ""])
    ws.append(["Google", "gemini-3.1-pro-preview", "", "tokens<=200000", 2, 12, "USD", "美元/百万tokens", "可使用", "代理商1", ""])
    ws.append(["Google", "gemini-3.1-pro-preview", "", "tokens>200000", 4, 18, "USD", "美元/百万tokens", "可使用", "代理商1", ""])
    ws.append(["DeepSeek(深度求索)", "deepseek-chat", "", "-", 2, 3, "CNY", "元/百万Token", "可使用", "自有", ""])
    ws.append(["Kling(可灵)", "kling-v2.5-turbo", "国内", "-", None, 0.7, "CNY", "元/秒", "未激活", "", "按秒计费示例"])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=model_import_template.xlsx"},
    )


@router.post("/import")
async def import_models(
    file: UploadFile = File(...),
    dry_run: bool = Query(False),
    db: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """批量导入模型基础列表（含阶梯计价合并）。source=manual_import。

    同名模型（name 相同）多档合并进 price_tiers JSON；单档直接写 input_price/output_price。
    - version_id：直接取「模型名称」列（人工导入场景下该列本身就是 API 调用 ID，如 deepseek-chat）。
    - upsert 键：优先按 version_id 匹配（能与 OpenRouter/Provider API 自动采集的同一模型合并，
      避免出现重复行）；未命中再按 name 兜底匹配（兼容历史手动导入数据）。
    - 地域（region）：「地域」列可手动填「国内/国际」覆盖；留空则按厂商/模型名关键词自动推断
      （阿里/字节/腾讯/百度/深度求索/月之暗面/智谱/MiniMax/可灵/零一万物/百川 → 国内）。
    """
    if not _is_admin(current_user):
        raise HTTPException(status_code=403, detail="仅管理员可操作")
    import pandas as pd
    import json as _json
    import math

    content = await file.read()
    try:
        df = pd.read_excel(BytesIO(content), sheet_name="模型基础列表")
    except Exception as e:
        raise HTTPException(400, f"读取 Excel 失败：{e}")

    def _clean(v):
        if v is None:
            return None
        if isinstance(v, float) and math.isnan(v):
            return None
        return v

    # 按模型名分组聚合档位
    model_groups: dict[str, dict] = {}
    for idx, row in df.iterrows():
        raw = {k: _clean(v) for k, v in row.items()}
        name = str(raw.get("模型名称") or "").strip()
        if not name:
            continue
        tier_label = str(raw.get("阶梯计价") or "-").strip()
        input_p = _clean(raw.get("输入价格"))
        output_p = _clean(raw.get("输出价格"))
        try:
            input_p = float(input_p) if input_p is not None else None
            output_p = float(output_p) if output_p is not None else None
        except (ValueError, TypeError):
            input_p = output_p = None

        if name not in model_groups:
            provider = str(raw.get("模型厂商") or "").strip() or None
            region_override = str(raw.get("地域") or "").strip()
            region = _normalize_region(region_override) if region_override else infer_region_from_text(provider, name)
            model_groups[name] = {
                "name": name,
                "provider": provider,
                "region": region,
                "price_currency": str(raw.get("价格币种") or "USD").strip(),
                "price_unit": str(raw.get("价格单位") or "美元/百万tokens").strip(),
                "is_active": _STATUS_MAP.get(str(raw.get("模型状态") or "").strip(), True),
                "suppliers_list": str(raw.get("供应商") or "").strip() or None,
                "description": str(raw.get("备注") or "").strip() or None,
                "input_price": None,
                "output_price": None,
                "tiers": [],
            }

        if tier_label != "-" and tier_label:
            model_groups[name]["tiers"].append({
                "threshold": tier_label,
                "input": input_p,
                "output": output_p,
            })
        else:
            model_groups[name]["input_price"] = input_p
            model_groups[name]["output_price"] = output_p

    stats = {"models": {"create": 0, "update": 0}}
    errors: list[dict] = []

    for idx, (name, data) in enumerate(model_groups.items()):
        tiers = data.pop("tiers")
        price_tiers_json: str | None = None
        if tiers:
            if not data["input_price"]:
                data["input_price"] = tiers[0].get("input")
                data["output_price"] = tiers[0].get("output")
            price_tiers_json = _json.dumps(tiers, ensure_ascii=False)

        try:
            # 优先按 version_id 匹配，能与自动采集（OpenRouter/Provider API）的同一模型合并，
            # 避免手动导入与自动采集各生成一行；未命中再按 name 兜底（兼容历史手动导入数据）
            existing = db.exec(select(ModelCatalog).where(ModelCatalog.version_id == name)).first()
            if not existing:
                existing = db.exec(select(ModelCatalog).where(ModelCatalog.name == name)).first()
            if existing:
                for field in ("provider", "region", "price_currency", "price_unit", "is_active",
                              "suppliers_list", "description", "input_price", "output_price"):
                    val = data.get(field)
                    if val is not None:
                        setattr(existing, field, val)
                existing.version_id = existing.version_id or name
                existing.price_tiers = price_tiers_json
                existing.source = "manual_import"
                if not dry_run:
                    db.add(existing)
                stats["models"]["update"] += 1
            else:
                new_m = ModelCatalog(
                    name=name,
                    version_id=name,
                    region=data.get("region", "international"),
                    source="manual_import",
                    is_active=data.get("is_active", True),
                    provider=data.get("provider"),
                    price_currency=data.get("price_currency", "USD"),
                    price_unit=data.get("price_unit", "美元/百万tokens"),
                    input_price=data.get("input_price"),
                    output_price=data.get("output_price"),
                    suppliers_list=data.get("suppliers_list"),
                    description=data.get("description"),
                    price_tiers=price_tiers_json,
                )
                if not dry_run:
                    db.add(new_m)
                stats["models"]["create"] += 1
        except Exception as e:
            errors.append({"sheet": "模型基础列表", "model": name, "reason": str(e)})

    if not dry_run:
        db.commit()

    return {
        "dry_run": dry_run,
        "stats": stats,
        "errors": errors,
        "message": "预检完成，未写入数据库" if dry_run else "导入完成",
    }
